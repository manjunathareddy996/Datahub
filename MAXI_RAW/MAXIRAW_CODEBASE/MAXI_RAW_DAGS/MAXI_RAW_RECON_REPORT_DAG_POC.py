import json
import logging
import smtplib
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from airflow import DAG
from airflow.models import Variable
from airflow.operators.python_operator import PythonOperator
from airflow.operators.trigger_dagrun import TriggerDagRunOperator
from airflow.providers.snowflake.hooks.snowflake import SnowflakeHook

# ---------------------------------------
# Configuration & Constants
# ---------------------------------------
IST = pytz.timezone('Asia/Kolkata')
CONFIG_JSON_PATH = "/opt/airflow/dags/Config/maxiraw_recon_queries.json"
SENDER_EMAIL = "DPM_L1_Team@bajajallianz.co.in"
SNOWFLAKE_CONN_ID = 'copy_cmd_snowflake'
RECIPIENT_EMAILS = [
    "ashraf.shaik01@bajajgeneral.com", "pratik.kulkarni@bajajgeneral.com", "satish.phirke@bajajgeneral.com"]

# ---------------------------------------
# DAG Definition
# ---------------------------------------
default_args = {
    "owner": "empower",
    "start_date": datetime(2025, 1, 1),
    "retries": 0,
    "retry_delay": timedelta(seconds=30),
    "provide_context": True,
}

dag = DAG(
    dag_id="MAXI_RAW_RECON_REPORT_DAG",
    default_args=default_args,
    description="Maximus Raw Reconciliation Report",
    schedule_interval=None,
    catchup=False,
    tags=["MAXIRAW", "RECON", "REPORTING"],
)


# ---------------------------------------
# Dynamic Label Parser
# ---------------------------------------
def parse_label_info(label: str) -> Dict[str, str]:
    """
    Dynamically extracts LOB, Entity, and Topic from label string.
    
    Examples:
    - "Travel Claims (NONHLTH_TRV_CLM) - Present in Maximus Raw Tables but Missing in MV_CLAIM_REGISTER"
      → LOB: "Travel", Entity: "Claim", Topic: "NONHLTH_TRV_CLM"
    - "Business Partners - Present in Maximus Raw Tables but Missing in ODS_PARTNER_DIM"
      → LOB: "Partner", Entity: "Customer", Topic: "BUSINESS_PARTNERS"
    """
    info = {
        'lob': '',
        'entity': '',
        'topic': ''
    }
    
    # Extract topic from parentheses (e.g., "NONHLTH_TRV_CLM")
    topic_match = re.search(r'\(([A-Z_]+)\)', label)
    if topic_match:
        info['topic'] = topic_match.group(1)
    
    # Extract LOB and Entity from the beginning of the label
    if topic_match:
        # Get text before the parentheses
        text_before_topic = label[:topic_match.start()].strip()
    else:
        # No parentheses, use the text before " - "
        text_before_topic = label.split(' - ')[0].strip()
    
    # Determine Entity type
    if 'Claims' in text_before_topic or 'Claim' in text_before_topic:
        info['entity'] = 'Claim'
        # Remove "Claims" or "Claim" to get LOB
        info['lob'] = text_before_topic.replace('Claims', '').replace('Claim', '').strip()
    elif 'Policy' in text_before_topic or 'Policies' in text_before_topic:
        info['entity'] = 'Policy'
        # Remove "Policy" or "Policies" to get LOB
        info['lob'] = text_before_topic.replace('Policy', '').replace('Policies', '').strip()
    elif 'Business Partners' in text_before_topic:
        info['entity'] = 'Customer'
        info['lob'] = 'Partner'
        info['topic'] = 'BUSINESS_PARTNERS'
    else:
        # Default case - treat entire text as LOB
        info['lob'] = text_before_topic
        info['entity'] = 'Unknown'
    
    return info


def determine_direction(label: str) -> str:
    """
    Determines if the discrepancy is "Missing in Register/ODS" or "Missing in JSON".
    """
    label_lower = label.lower()
    
    if 'missing in mv_' in label_lower or 'missing in ods_' in label_lower:
        return 'MISSING_IN_REGISTER'
    elif 'missing in maximus raw' in label_lower or 'present in mv_' in label_lower or 'present in ods_' in label_lower:
        return 'MISSING_IN_JSON'
    else:
        return 'UNKNOWN'


def extract_raw_table_from_sql(sql: str) -> str:
    """
    Extracts the raw table name from SQL query.
    
    Example:
    SQL: "SELECT CLAIMNO FROM BAGIC_PROD_MIRROR_DB.MAXI_RAW.NONHLTH_TRV_CLM_VW_DATA_CLAIM_DETAILS A,..."
    Returns: "BAGIC_PROD_MIRROR_DB.MAXI_RAW.NONHLTH_TRV_CLM_VW_DATA_CLAIM_DETAILS"
    """
    # Look for pattern: FROM [schema].[table_name]
    match = re.search(r'FROM\s+(BAGIC_PROD_MIRROR_DB\.MAXI_RAW\.\w+)', sql, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def get_staging_table_name(topic: str) -> str:
    """
    Converts topic name to staging table name.
    
    Example:
    - "NONHLTH_TRV_CLM" -> "BAGIC_PROD_STAGING_DB.MAXI_RAW.MAXI_SUR_BND_UWR"
    - "MAXI_SUR_BND_UWR" -> "BAGIC_PROD_STAGING_DB.MAXI_RAW.MAXI_SUR_BND_UWR"
    """
    if not topic:
        return None
    
    # The staging table follows pattern: BAGIC_PROD_STAGING_DB.MAXI_RAW.[TOPIC]
    return f"BAGIC_PROD_STAGING_DB.MAXI_RAW.{topic}"


# ---------------------------------------
# Excel Summary Report Generation
# ---------------------------------------
def generate_excel_summary(recon_results: List[Dict[str, Any]], total_counts: Dict[str, int], report_date: str, hook: SnowflakeHook, max_timestamps: Dict[str, str]) -> str:
    """
    Generates an Excel summary report showing all LOBs and their reconciliation status.
    FULLY DYNAMIC - extracts all information from labels and config.
    Returns the file path of the generated Excel file.
    """
    excel_path = f"/tmp/MAXI_RAW_RECON_SUMMARY_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Prepare summary data dynamically
    summary_data = []
    sr_no = 1
    
    # Group results by LOB + Entity + Topic to combine "Missing in Register" and "Missing in JSON" rows
    grouped_results = {}
    
    for result in recon_results:
        label = result['label']
        status = result['status']
        issue = result.get('issue', '')
        sql = result.get('sql', '')
        
        # Parse label to extract LOB, Entity, Topic
        label_info = parse_label_info(label)
        lob = label_info['lob']
        entity = label_info['entity']
        topic = label_info['topic']
        
        # Determine direction
        direction = determine_direction(label)
        
        # Create a unique key for grouping
        group_key = f"{lob}|{entity}|{topic}"
        
        # Initialize group if not exists
        if group_key not in grouped_results:
            grouped_results[group_key] = {
                'lob': lob,
                'entity': entity,
                'topic': topic,
                'issue': issue if issue else '',
                'total_count_json': total_counts.get(topic, 0),  # Use pre-calculated total count
                'missing_in_register': 0,
                'missing_in_json': 0,
                'sql': sql
            }
        
        # Update issue if present (take the first non-empty issue)
        if issue and not grouped_results[group_key]['issue']:
            grouped_results[group_key]['issue'] = issue
        
        # Store SQL for reference
        if sql and not grouped_results[group_key]['sql']:
            grouped_results[group_key]['sql'] = sql
        
        # Update counts based on direction and status
        if status == 'success':
            record_count = result.get('record_count', 0)
            
            if direction == 'MISSING_IN_REGISTER':
                grouped_results[group_key]['missing_in_register'] = record_count
            elif direction == 'MISSING_IN_JSON':
                grouped_results[group_key]['missing_in_json'] = record_count
        elif status == 'issue':
            # For issues, mark as such but don't add counts
            pass
    
    # Convert grouped results to list
    for group_key, group_data in grouped_results.items():
        summary_data.append({
            'Sr No': sr_no,
            'LOB': group_data['lob'],
            'Entity': group_data['entity'],
            'TOPIC': group_data['topic'],
            'ISSUE': group_data['issue'],
            'TOTAL COUNT IN JSON': group_data['total_count_json'],
            'MAX FILE TIMESTAMP OF JSON': max_timestamps.get(group_data['topic'], 'N/A'),
            'MISSING IN REGISTER BUT PRESENT IN JSON': group_data['missing_in_register'],
            'MISSING IN JSON BUT PRESENT IN REGISTER': group_data['missing_in_json']
        })
        sr_no += 1
    
    # Sort by LOB, then Entity
    summary_data.sort(key=lambda x: (x['LOB'], x['Entity']))
    
    # Re-number after sorting
    for idx, row in enumerate(summary_data, 1):
        row['Sr No'] = idx
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation Summary"
    
    # Define headers (added MAX FILE TIMESTAMP OF JSON)
    headers = ['Sr No', 'LOB', 'Entity', 'TOPIC', 'ISSUE', 
               'TOTAL COUNT IN JSON', 'MAX FILE TIMESTAMP OF JSON', 'MISSING IN REGISTER BUT PRESENT IN JSON', 'MISSING IN JSON BUT PRESENT IN REGISTER']
    
    # Header styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    data_font = Font(name='Arial', size=10)
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row_num, data_row in enumerate(summary_data, 2):
        # Alternate row colors
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = data_row.get(header, '')
            
            # Handle "NO" display for ISSUE column when empty
            if header == 'ISSUE' and not value:
                cell.value = 'NO'
            else:
                cell.value = value
            
            cell.font = data_font
            cell.border = border
            cell.fill = row_fill
            
            # Center align numbers and short text, left align long text
            if header in ['Sr No', 'LOB', 'Entity', 'TOTAL COUNT IN JSON', 
                         'MISSING IN REGISTER BUT PRESENT IN JSON', 'MISSING IN JSON BUT PRESENT IN REGISTER']:
                cell.alignment = data_alignment_center
            else:
                cell.alignment = data_alignment_left
            
            # Highlight missing counts in red if > 0
            if header in ['MISSING IN REGISTER BUT PRESENT IN JSON', 'MISSING IN JSON BUT PRESENT IN REGISTER']:
                if isinstance(value, (int, float)) and value > 0:
                    cell.font = Font(name='Arial', size=10, color='D13438', bold=True)
    
    # Set column widths (added MAX FILE TIMESTAMP OF JSON column)
    column_widths = {
        'A': 8,   # Sr No
        'B': 15,  # LOB
        'C': 12,  # Entity
        'D': 25,  # TOPIC
        'E': 50,  # ISSUE
        'F': 20,  # TOTAL COUNT IN JSON
        'G': 30,  # MAX FILE TIMESTAMP OF JSON
        'H': 35,  # MISSING IN REGISTER BUT PRESENT IN JSON
        'I': 35   # MISSING IN JSON BUT PRESENT IN REGISTER
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(excel_path)
    
    logging.info(f"✓ Excel summary report generated: {excel_path}")
    logging.info(f"  Total unique LOB+Entity combinations: {len(summary_data)}")
    
    return excel_path


# ---------------------------------------
# HTML Template Generation
# ---------------------------------------
def generate_html_report(recon_results: List[Dict[str, Any]], report_date: str) -> str:

    # Calculate summary statistics
    total_queries = len(recon_results)
    successful_queries = sum(1 for r in recon_results if r['status'] == 'success')
    failed_queries = total_queries - successful_queries
    total_records = sum(r.get('record_count', 0) for r in recon_results if r['status'] == 'success')
    
    # Generate sections HTML
    sections_html = []
    for idx, result in enumerate(recon_results, 1):
        section_html = generate_section_html(result, idx)
        sections_html.append(section_html)
    
    sections_content = "\n".join(sections_html)
    
    # Main HTML template with inline styles only
    html_template = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
</head>
<body style="margin:0; padding:20px; font-family:Arial,sans-serif; background-color:#f5f5f5;">
    <table width="100%" cellpadding="0" cellspacing="0" style="max-width:1200px; margin:0 auto; background-color:#ffffff; border:1px solid #d4d4d4;">
        <!-- Header -->
        <tr>
            <td style="background-color:#0078d4; color:#ffffff; padding:30px; text-align:left;">
                <h1 style="margin:0 0 10px 0; font-size:28px; font-weight:600;">📊 Maximus Raw Reconciliation Report</h1>
                <div style="display:inline-block; padding:8px 16px; background-color:rgba(255,255,255,0.2); border:1px solid rgba(255,255,255,0.3); border-radius:4px; font-size:16px;">
                    🗓️ Generated on {report_date}
                </div>
            </td>
        </tr>
        
        <!-- Summary Cards -->
        <tr>
            <td style="padding:0;">
                <table width="100%" cellpadding="0" cellspacing="0" border="0">
                    <tr>
                        <!-- Total Queries -->
                        <td width="25%" style="padding:24px; text-align:center; border-right:1px solid #d4d4d4; border-bottom:1px solid #d4d4d4; vertical-align:top;">
                            <div style="font-size:40px; margin-bottom:12px;">📋</div>
                            <div style="font-size:12px; color:#666666; font-weight:600; text-transform:uppercase; margin-bottom:8px;">TOTAL QUERIES</div>
                            <div style="font-size:36px; font-weight:700; color:#0078d4;">{total_queries}</div>
                        </td>
                        
                        <!-- Successful -->
                        <td width="25%" style="padding:24px; text-align:center; border-right:1px solid #d4d4d4; border-bottom:1px solid #d4d4d4; vertical-align:top;">
                            <div style="font-size:40px; margin-bottom:12px; color:#107c10;">✓</div>
                            <div style="font-size:12px; color:#666666; font-weight:600; text-transform:uppercase; margin-bottom:8px;">SUCCESSFUL</div>
                            <div style="font-size:36px; font-weight:700; color:#107c10;">{successful_queries}</div>
                        </td>
                        
                        <!-- Failed -->
                        <td width="25%" style="padding:24px; text-align:center; border-right:1px solid #d4d4d4; border-bottom:1px solid #d4d4d4; vertical-align:top;">
                            <div style="font-size:40px; margin-bottom:12px; color:#d13438;">✗</div>
                            <div style="font-size:12px; color:#666666; font-weight:600; text-transform:uppercase; margin-bottom:8px;">FAILED</div>
                            <div style="font-size:36px; font-weight:700; color:#d13438;">{failed_queries}</div>
                        </td>
                        
                        <!-- Total Records -->
                        <td width="25%" style="padding:24px; text-align:center; border-bottom:1px solid #d4d4d4; vertical-align:top;">
                            <div style="font-size:40px; margin-bottom:12px;">∑</div>
                            <div style="font-size:12px; color:#666666; font-weight:600; text-transform:uppercase; margin-bottom:8px;">TOTAL RECORDS</div>
                            <div style="font-size:36px; font-weight:700; color:#0078d4;">{total_records:,}</div>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        
        <!-- Main Content -->
        <tr>
            <td style="padding:30px;">
                {sections_content}
            </td>
        </tr>
        
        <!-- Footer -->
        <tr>
            <td style="background-color:#0078d4; color:#ffffff; padding:25px; text-align:center;">
                <p style="margin:0 0 10px 0; font-size:14px; font-weight:600;">Bajaj General Insurance</p>
                <div style="height:1px; background-color:rgba(255,255,255,0.3); margin:15px auto; max-width:600px;"></div>
                <p style="margin:0; font-size:12px; font-style:italic;">
                    This is an automated report. For queries or issues, please contact the DPM team.<br>
                    📎 Please find the detailed Excel summary report attached.
                </p>
            </td>
        </tr>
    </table>
</body>
</html>
"""
    
    return html_template


def generate_section_html(result: Dict[str, Any], section_number: int) -> str:

    label = result['label']
    status = result['status']
    df = result.get('dataframe')
    error = result.get('error')
    issue = result.get('issue')
    total_record_count = result.get('record_count', 0)
    
    if status == 'error':
        return f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px; border:1px solid #d4d4d4;">
            <tr>
                <td style="background-color:#d13438; color:#ffffff; padding:16px;">
                    <table width="100%" cellpadding="0" cellspacing="0">
                        <tr>
                            <td style="color:#ffffff;">
                                <span style="display:inline-block; padding:6px 10px; background-color:rgba(255,255,255,0.25); border-radius:3px; font-weight:700; margin-right:10px;">{section_number}</span>
                                <span style="font-size:16px; font-weight:600;">{label}</span>
                            </td>
                            <td align="right" style="color:#ffffff;">
                                <span style="background-color:rgba(255,255,255,0.25); padding:6px 14px; border-radius:20px; font-size:13px; font-weight:600;">✗ ERROR</span>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            <tr>
                <td style="padding:20px; background-color:#fde7e9; border-left:4px solid #d13438; color:#a80000;">
                    <strong style="display:block; margin-bottom:8px;">⚠ Query Execution Failed:</strong>
                    {error}
                </td>
            </tr>
        </table>
        """
    
    if status == 'issue':
        return f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px; border:1px solid #d4d4d4;">
            <tr>
                <td style="background-color:#d13438; color:#ffffff; padding:16px;">
                    <table width="100%" cellpadding="0" cellspacing="0">
                        <tr>
                            <td style="color:#ffffff;">
                                <span style="display:inline-block; padding:6px 10px; background-color:rgba(255,255,255,0.25); border-radius:3px; font-weight:700; margin-right:10px;">{section_number}</span>
                                <span style="font-size:16px; font-weight:600;">{label}</span>
                            </td>
                            <td align="right" style="color:#ffffff;">
                                <span style="background-color:rgba(255,255,255,0.25); padding:6px 14px; border-radius:20px; font-size:13px; font-weight:600;">! ISSUE</span>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            <tr>
                <td style="padding:20px; background-color:#fde7e9; border-left:4px solid #d13438; color:#a80000;">
                    <strong style="display:block; margin-bottom:8px;">⚠ Issue in Data:</strong>
                    {issue}
                </td>
            </tr>
        </table>
        """
    
    # Success case
    if total_record_count == 0:
        return f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px; border:1px solid #d4d4d4;">
            <tr>
                <td style="background-color:#0078d4; color:#ffffff; padding:16px;">
                    <table width="100%" cellpadding="0" cellspacing="0">
                        <tr>
                            <td style="color:#ffffff;">
                                <span style="display:inline-block; padding:6px 10px; background-color:rgba(255,255,255,0.25); border-radius:3px; font-weight:700; margin-right:10px;">{section_number}</span>
                                <span style="font-size:16px; font-weight:600;">{label}</span>
                            </td>
                            <td align="right" style="color:#ffffff;">
                                <span style="background-color:rgba(255,255,255,0.25); padding:6px 14px; border-radius:20px; font-size:13px; font-weight:600;">0 records</span>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            <tr>
                <td style="padding:50px 20px; text-align:center; background-color:#f9f9f9;">
                    <div style="font-size:48px; margin-bottom:12px; color:#107c10;">✓</div>
                    <div style="font-size:15px; font-weight:500; color:#666666;">No discrepancies found - All records reconciled successfully!</div>
                </td>
            </tr>
        </table>
        """
    
    # Display top 5 records
    showing_top_5 = total_record_count > 5
    
    # Generate table HTML with proper formatting
    table_html = generate_table_html(df)
    
    # Add message if showing only top 5
    showing_message = ""
    if showing_top_5:
        remaining = total_record_count - 5
        showing_message = f"""
            <tr>
                <td style="padding:12px 20px; background-color:#fff4ce; border-top:1px solid #d4d4d4; color:#8a6d3b; font-size:13px; text-align:center;">
                    <span style="margin-right:6px;">ⓘ</span>
                    Showing top 5 records out of {total_record_count:,} total records ({remaining:,} more records not displayed)
                </td>
            </tr>
        """
    
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px; border:1px solid #d4d4d4;">
        <tr>
            <td style="background-color:#0078d4; color:#ffffff; padding:16px;">
                <table width="100%" cellpadding="0" cellspacing="0">
                    <tr>
                        <td style="color:#ffffff;">
                            <span style="display:inline-block; padding:6px 10px; background-color:rgba(255,255,255,0.25); border-radius:3px; font-weight:700; margin-right:10px;">{section_number}</span>
                            <span style="font-size:16px; font-weight:600;">{label}</span>
                        </td>
                        <td align="right" style="color:#ffffff;">
                            <span style="background-color:rgba(255,255,255,0.25); padding:6px 14px; border-radius:20px; font-size:13px; font-weight:600;">{total_record_count:,} records</span>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        <tr>
            <td style="padding:20px;">
                {table_html}
            </td>
        </tr>
        {showing_message}
    </table>
    """


def generate_table_html(df: pd.DataFrame) -> str:

    if df.empty:
        return '<div style="padding:50px 20px; text-align:center; background-color:#f9f9f9;"><div style="font-size:48px; margin-bottom:12px;">📭</div><div style="font-size:15px; color:#666666;">No data available</div></div>'
    
    # Start table with inline styles
    html_parts = ['<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse; font-size:13px;">']
    
    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th style="padding:12px 16px; text-align:left; font-weight:600; color:#333333; border:1px solid #d4d4d4; text-transform:uppercase; font-size:12px; background-color:#f8f8f8;">{col}</th>')
    html_parts.append('</tr></thead>')
    
    # Table body
    html_parts.append('<tbody>')
    for row_idx, row in df.iterrows():
        # Alternate row colors
        bg_color = '#fafafa' if row_idx % 2 == 1 else '#ffffff'
        html_parts.append('<tr>')
        for col in df.columns:
            value = row[col]
            
            # Check if value is numeric for right alignment
            is_numeric = pd.api.types.is_numeric_dtype(type(value)) or isinstance(value, (int, float))
            
            # Format the value
            if pd.isna(value):
                formatted_value = '<em style="color:#999999;">NULL</em>'
                text_align = 'left'
            elif is_numeric:
                if isinstance(value, float):
                    formatted_value = f'{value:,.2f}'
                else:
                    formatted_value = f'{value:,}'
                text_align = 'right'
            else:
                formatted_value = str(value)
                text_align = 'left'
            
            html_parts.append(f'<td style="padding:12px 16px; color:#333333; border:1px solid #d4d4d4; text-align:{text_align}; background-color:{bg_color};">{formatted_value}</td>')
        html_parts.append('</tr>')
    html_parts.append('</tbody>')
    
    html_parts.append('</table>')
    
    return ''.join(html_parts)


# ---------------------------------------
# Email Functions
# ---------------------------------------
def send_email(subject: str, body_html: str, excel_path: str = None, recipients: List[str] = None):
    """
    Send email with optional Excel attachment.
    """
    if recipients is None:
        recipients = RECIPIENT_EMAILS
    
    try:
        # Get SMTP settings from Airflow Variables
        smtp_settings_json = Variable.get("smtp_settings_password")
        smtp_settings = json.loads(smtp_settings_json)
        
        smtp_host = smtp_settings.get("smtp_host")
        smtp_port = int(smtp_settings.get("smtp_port"))
        smtp_user = smtp_settings.get("smtp_user")
        smtp_password = smtp_settings.get("smtp_password")
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg["From"] = SENDER_EMAIL
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        
        # Attach HTML content
        html_part = MIMEText(body_html, "html", "utf-8")
        msg.attach(html_part)
        
        # Attach Excel file if provided
        if excel_path:
            try:
                with open(excel_path, 'rb') as f:
                    excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    excel_attachment.set_payload(f.read())
                    encoders.encode_base64(excel_attachment)
                    
                    filename = excel_path.split('/')[-1]
                    excel_attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                    msg.attach(excel_attachment)
                    
                logging.info(f"  ✓ Excel attachment added: {filename}")
            except Exception as e:
                logging.error(f"  ✗ Failed to attach Excel file: {e}")
        
        # Send email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        
        logging.info(f"✓ Email sent successfully: {subject}")
        logging.info(f"  Recipients: {len(recipients)} addresses")
        
    except Exception as e:
        logging.error(f"✗ Failed to send email: {e}")
        raise


# ---------------------------------------
# Calculate Total Counts and Max File Timestamps from Raw Tables
# ---------------------------------------
def calculate_total_counts(recon_queries: List[Dict[str, Any]], hook: SnowflakeHook) -> Dict[str, int]:
    """
    Calculate total counts from raw tables for each TOPIC.
    Returns a dictionary mapping TOPIC -> total_count
    """
    total_counts = {}
    processed_topics = set()
    
    for query_config in recon_queries:
        label = query_config.get("label", "")
        sql = query_config.get("sql", "")
        
        # Parse label to get topic
        label_info = parse_label_info(label)
        topic = label_info['topic']
        
        # Skip if already processed this topic
        if topic in processed_topics or not topic:
            continue
        
        # Extract raw table from SQL
        raw_table = extract_raw_table_from_sql(sql)
        
        if raw_table:
            try:
                count_sql = f"SELECT COUNT(*) as TOTAL_COUNT FROM {raw_table}"
                logging.info(f"  Calculating total count for {topic} from {raw_table}")
                
                count_df = hook.get_pandas_df(count_sql)
                total_count = int(count_df['TOTAL_COUNT'].iloc[0])
                
                total_counts[topic] = total_count
                processed_topics.add(topic)
                
                logging.info(f"  ✓ {topic}: {total_count:,} records")
            except Exception as e:
                logging.error(f"  ✗ Failed to get count for {topic}: {e}")
                total_counts[topic] = 0
        else:
            logging.warning(f"  ⚠ Could not extract raw table from SQL for {topic}")
            total_counts[topic] = 0
    
    return total_counts


def calculate_max_file_timestamps(recon_queries: List[Dict[str, Any]], hook: SnowflakeHook) -> Dict[str, str]:
    """
    Calculate max file_timestamp from staging tables for each TOPIC.
    Returns a dictionary mapping TOPIC -> max_file_timestamp
    """
    max_timestamps = {}
    processed_topics = set()
    
    for query_config in recon_queries:
        label = query_config.get("label", "")
        
        # Parse label to get topic
        label_info = parse_label_info(label)
        topic = label_info['topic']
        
        # Skip if already processed this topic
        if topic in processed_topics or not topic:
            continue
        
        # Get staging table name
        staging_table = get_staging_table_name(topic)
        
        if staging_table:
            try:
                max_timestamp_sql = f"SELECT MAX(file_timestamp) as MAX_FILE_TIMESTAMP FROM {staging_table}"
                logging.info(f"  Calculating max file_timestamp for {topic} from {staging_table}")
                
                timestamp_df = hook.get_pandas_df(max_timestamp_sql)
                max_timestamp = timestamp_df['MAX_FILE_TIMESTAMP'].iloc[0]
                
                # Convert to string if not None
                if pd.notna(max_timestamp):
                    max_timestamps[topic] = str(max_timestamp)
                else:
                    max_timestamps[topic] = 'N/A'
                
                processed_topics.add(topic)
                
                logging.info(f"  ✓ {topic}: {max_timestamps[topic]}")
            except Exception as e:
                logging.error(f"  ✗ Failed to get max file_timestamp for {topic}: {e}")
                max_timestamps[topic] = 'Error'
        else:
            logging.warning(f"  ⚠ Could not determine staging table for {topic}")
            max_timestamps[topic] = 'N/A'
    
    return max_timestamps


# ---------------------------------------
# Main Reconciliation Task
# ---------------------------------------
def run_reconciliation_report(**kwargs):
    """
    Executes reconciliation queries and generates beautiful HTML report with Excel summary.
    FULLY DYNAMIC - extracts all information from config labels.
    OPTIMIZED: Fetches count first, then only top 5 records.
    """
    logging.info("=" * 80)
    logging.info("Starting Premium Reconciliation Report Generation (Fully Dynamic)")
    logging.info("=" * 80)
    
    # Load configuration
    try:
        with open(CONFIG_JSON_PATH, 'r') as f:
            recon_queries = json.load(f)
        logging.info(f"✓ Loaded {len(recon_queries)} reconciliation queries from config")
    except Exception as e:
        logging.error(f"✗ Failed to load config file: {e}")
        raise
    
    # Initialize Snowflake connection
    try:
        hook = SnowflakeHook(snowflake_conn_id=SNOWFLAKE_CONN_ID)
        logging.info("✓ Connected to Snowflake")
    except Exception as e:
        logging.error(f"✗ Failed to connect to Snowflake: {e}")
        raise
    
    # Calculate total counts from raw tables
    logging.info("\n" + "=" * 80)
    logging.info("Calculating Total Counts from Raw Tables")
    logging.info("=" * 80)
    
    total_counts = calculate_total_counts(recon_queries, hook)
    
    # Calculate max file timestamps from staging tables
    logging.info("\n" + "=" * 80)
    logging.info("Calculating Max File Timestamps from Staging Tables")
    logging.info("=" * 80)
    
    max_timestamps = calculate_max_file_timestamps(recon_queries, hook)
    
    # Execute queries and collect results
    logging.info("\n" + "=" * 80)
    logging.info("Executing Reconciliation Queries")
    logging.info("=" * 80)
    
    recon_results = []
    
    for idx, query_config in enumerate(recon_queries, 1):
        label = query_config.get("label", f"Query_{idx}")
        sql = query_config.get("sql", "")
        issue = query_config.get("issue", "")
        
        logging.info(f"\n[{idx}/{len(recon_queries)}] Executing: {label}")
        logging.info("-" * 60)
        
        if not sql:
            logging.warning(f"  ⚠ Empty SQL query for {label}, skipping...")
            recon_results.append({
                'label': label,
                'status': 'error',
                'error': 'Empty SQL query provided',
                'dataframe': pd.DataFrame(),
                'record_count': 0,
                'issue': issue,
                'sql': sql
            })
            continue

        if issue:
            logging.warning(f"  ⚠ issue for {label}, skipping...")
            recon_results.append({
                'label': label,
                'status': 'issue',
                'issue': issue,
                'dataframe': pd.DataFrame(),
                'sql': sql
            })
            continue
        
        try:
            # OPTIMIZATION 1: First get the count
            count_sql = f"SELECT COUNT(*) as TOTAL_COUNT FROM ({sql})"
            logging.info(f"  → Fetching count...")
            
            count_df = hook.get_pandas_df(count_sql)
            record_count = int(count_df['TOTAL_COUNT'].iloc[0])
            
            logging.info(f"  ✓ Total records found: {record_count:,}")
            
            # OPTIMIZATION 2: Only fetch top 5 records if count > 0
            if record_count > 0:
                limited_sql = f"SELECT * FROM ({sql}) LIMIT 5"
                logging.info(f"  → Fetching top 5 records...")
                df = hook.get_pandas_df(limited_sql)
                logging.info(f"  ✓ Retrieved {len(df)} sample records")
            else:
                df = pd.DataFrame()
                logging.info(f"  ✓ No records to fetch")
            
            recon_results.append({
                'label': label,
                'status': 'success',
                'dataframe': df,
                'record_count': record_count,
                'issue': issue,
                'sql': sql
            })
            
        except Exception as e:
            error_msg = str(e)
            logging.error(f"  ✗ Query failed: {error_msg}")
            
            recon_results.append({
                'label': label,
                'status': 'error',
                'error': error_msg,
                'dataframe': pd.DataFrame(),
                'record_count': 0,
                'issue': issue,
                'sql': sql
            })
    
    # Generate report date
    report_date = datetime.now(IST).strftime('%B %d, %Y at %I:%M %p IST')
    
    # Generate Excel summary report (FULLY DYNAMIC)
    logging.info("\n" + "=" * 80)
    logging.info("Generating Dynamic Excel Summary Report")
    logging.info("=" * 80)
    
    excel_path = generate_excel_summary(recon_results, total_counts, report_date, hook, max_timestamps)
    
    # Generate HTML report
    logging.info("\n" + "=" * 80)
    logging.info("Generating HTML Report")
    logging.info("=" * 80)
    
    html_body = generate_html_report(recon_results, report_date)
    
    # Send email with Excel attachment
    email_subject = f"📊 Maximus Raw Reconciliation Report - {datetime.now(IST).strftime('%Y-%m-%d')}"
    
    logging.info(f"\nSending email: {email_subject}")
    send_email(email_subject, html_body, excel_path)
    
    # Summary
    successful = sum(1 for r in recon_results if r['status'] == 'success')
    failed = len(recon_results) - successful
    total_records = sum(r.get('record_count', 0) for r in recon_results if r['status'] == 'success')
    
    logging.info("\n" + "=" * 80)
    logging.info("RECONCILIATION REPORT SUMMARY")
    logging.info("=" * 80)
    logging.info(f"Total Queries:      {len(recon_queries)}")
    logging.info(f"Successful:         {successful}")
    logging.info(f"Failed:             {failed}")
    logging.info(f"Total Records:      {total_records:,}")
    logging.info(f"Excel Report:       {excel_path}")
    logging.info("=" * 80)
    logging.info("✓ Report generation completed successfully!")
    logging.info("=" * 80 + "\n")


# ---------------------------------------
# DAG Tasks
# ---------------------------------------
run_recon_task = PythonOperator(
    task_id="run_reconciliation_report",
    python_callable=run_reconciliation_report,
    dag=dag,
)


run_recon_task
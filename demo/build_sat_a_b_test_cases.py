"""Generate sat_a_b_test_cases.xlsx — all test cases for sat_a_b.sql (sat_multi_source macro).

Columns: test_case_no | test_case_desc | actual_output | expected_output
actual_output is left blank for the tester to fill after running dbt.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "sat_a_b_test_cases"

headers = ["test_case_no", "test_case_desc", "actual_output", "expected_output"]

# (desc, expected_output)
rows = [
    # ---- Setup / delegation ----
    ("First run, sat table does not exist yet (fresh schema). Watermark falls back to sentinel 1900-01-01, full load.",
     "Table created. All in-window source rows loaded."),
    ("First run, single source only (TABLE_A seeded, TABLE_B empty).",
     "1 row: PARTY_HKEY, PHONE_1 populated, PHONE_2=NULL, RECORD_SOURCE=TABLE_A."),
    ("First run, both sources same PK (id=1001 in A and B).",
     "2 rows for the PK: one RECORD_SOURCE=TABLE_A (PHONE_2 NULL), one RECORD_SOURCE=TABLE_B (PHONE_1+PHONE_2). Independent hashdiff timelines."),
    ("source_model passed as a single string instead of a list.",
     "Delegates to automate_dv.sat(); output identical to a plain single-source satellite."),

    # ---- Core dedup / change detection ----
    ("Section 30 regression: TABLE_B loads first, then TABLE_A reloads UNCHANGED data with a newer updated_at.",
     "No duplicate. A compares against A's own last row (same hashdiff) -> suppressed. 0 new rows for TABLE_A."),
    ("TABLE_A genuinely changes PHONE_1; TABLE_B unchanged.",
     "1 new row for TABLE_A only. TABLE_B untouched."),
    ("Both sources change payload in the same run.",
     "2 new rows: one per source, each with its new hashdiff."),
    ("Unchanged reload of BOTH sources (newer updated_at, identical phone values).",
     "0 rows inserted. LAG suppresses both per (PK, SOURCE)."),
    ("Interleaved timestamps same PK: A@t1, B@t2, A@t3 (A changed between t1 and t3).",
     "3 rows: A@t1, B@t2, A@t3. Each source timeline evaluated independently; no phantom versions from interleaving."),
    ("Same source emits two consecutive rows with identical hashdiff (no real change).",
     "Only the first is kept; the identical follow-up is suppressed by LAG."),

    # ---- Column alignment / superset ----
    ("Column present in one source, absent in another (TABLE_A has no PHONE_2).",
     "TABLE_A rows show PHONE_2 = NULL (CAST(NULL AS VARCHAR)). No 'invalid identifier A.PHONE_2' error."),
    ("src_payload provided explicitly (authoritative superset).",
     "Superset = exactly the src_payload columns, sorted: PHONE_1, PHONE_2."),
    ("src_payload omitted but src_column_map provided (derived superset).",
     "Superset derived from union of mapped columns minus system columns; same column set as explicit payload."),
    ("src_column_map contains a model NOT in source_model (e.g. stg2_ghost).",
     "Compile-time WARNING logged ('...not in source_model list...Ignoring.'). Model still compiles; output unchanged."),
    ("src_column_map missing an entry for a source_model (drop stg2_a from map).",
     "That source gets [] -> every payload column padded NULL for its rows."),
    ("src_extra_columns provided as a string.",
     "Column merged into superset (deduped, sorted). Present in output."),
    ("src_extra_columns provided as a list.",
     "All listed columns merged into superset (deduped, sorted). Present in output."),
    ("Column casing mismatch (physical phone_1 vs superset PHONE_1).",
     "Case-insensitive match -> emits CAST(a.PHONE_1 AS VARCHAR); no duplicate NULL-padded column."),
    ("Payload values differ in type across sources (numeric vs masked string).",
     "All payload cast to VARCHAR in every UNION ALL branch; no 'Numeric value not recognized' error."),
    ("src_run_ts column name would collide with a payload column.",
     "src_run_ts is rejected from the payload superset; stamped once as a system column, never padded/duplicated."),

    # ---- Watermark ----
    ("Second incremental run: watermark = MAX(DBT_RUN_TS) from the sat.",
     "Only rows with LOAD_DATETIME > previous MAX(DBT_RUN_TS) and <= to_date are processed; old rows not reprocessed."),
    ("Explicit var('from_date') and var('to_date') override the window.",
     "Only rows with from_date < LOAD_DATETIME <= to_date flow through."),
    ("var('to_date') only (no from_date) with existing sat.",
     "from_date = MAX(DBT_RUN_TS); to_date = the override literal."),
    ("Run where no source rows fall in the watermark window.",
     "0 rows inserted; no error."),

    # ---- Data hygiene ----
    ("Source row with NULL src_pk (NULL id).",
     "Filtered out by WHERE PARTY_HKEY IS NOT NULL; no row emitted for it."),

    # ---- Validation / error paths ----
    ("Omit required param src_pk.", "Compiler error: 'src_pk is a required parameter for sat_multi_source'."),
    ("Omit required param src_hashdiff.", "Compiler error: 'src_hashdiff is a required parameter for sat_multi_source'."),
    ("Omit required param src_payload.", "Compiler error: 'src_payload is a required parameter for sat_multi_source'."),
    ("Omit required param src_ldts.", "Compiler error: 'src_ldts is a required parameter for sat_multi_source'."),
    ("Omit required param src_source.", "Compiler error: 'src_source is a required parameter for sat_multi_source'."),
    ("Omit required param source_model.", "Compiler error: 'source_model is a required parameter for sat_multi_source'."),
    ("source_model is an empty list [].", "Compiler error: 'source_model list must contain at least one model name'."),
    ("source_model list contains an empty/blank string entry.", "Compiler error: 'source_model entry at position N must be a non-empty string'."),
    ("source_model is a dict/mapping.", "Compiler error: 'source_model must be a string or a list of strings'."),
    ("Superset resolves to zero payload columns.", "Compiler error: 'No payload columns found across source models'."),
]

# Styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")

for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=1 + (c - 1), value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for i, (desc, expected) in enumerate(rows, start=1):
    r = i + 1
    ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=2, value=desc).alignment = wrap_top
    ws.cell(row=r, column=3, value="").alignment = wrap_top          # actual_output (tester fills)
    ws.cell(row=r, column=4, value=expected).alignment = wrap_top
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = border

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 70
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 60
ws.freeze_panes = "A2"

out = "sat_a_b_test_cases.xlsx"
wb.save(out)
print("Wrote", out, "with", len(rows), "test cases")
